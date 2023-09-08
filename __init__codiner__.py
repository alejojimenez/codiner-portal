from codigo.app_codiner import Scraper_Codiner
#import smtplib
from openpyxl import load_workbook
import datetime


def send_notification():
    # Código para enviar correo electrónico de notificación
    print('')

if __name__ == '__main__':
    
    print('Obteniendo credenciales...')
    print('----------------------------------------------------------------------')
        
    credencials = '.\\config\\credenciales.xlsx'
    libro_accesos = load_workbook(credencials)
    hoja_credenciales = libro_accesos['Hoja1']
        
    for j in hoja_credenciales.iter_rows(2):
        try:
            rut = j[0].value
            passw = j[1].value
            web = j[2].value
            break
        except:
            ('no hay credenciales')
            
    email = rut
    password = passw
    url = web
    driver_path = 'chromedriver.exe'
    
    scraper = Scraper_Codiner(url, email, password, driver_path)
    
    #Primer ingreso, año actual
    print('ingresamos en la clase Scraper_Codiner...')
    print('----------------------------------------------------------------------')
    
    scraper.login()
    print('hacemos login en el portal...')
    print('----------------------------------------------------------------------')
        
    scraper.scrapping_codiner()
    print('hacemos scrapping al portal...')
    print('----------------------------------------------------------------------')
    
    scraper.archivos()
    print('hacemos scrapping al portal...')
    print('----------------------------------------------------------------------')    
    
    scraper.close()
    print('cerramos el bot final...')
    print('----------------------------------------------------------------------')