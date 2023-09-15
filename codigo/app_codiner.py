import os
import time
import shutil
import requests
import pandas as pd
import re
import glob
import fitz
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.action_chains import ActionChains

class Scraper_Codiner():

    def __init__(self,url, email, password, driver_path):
        print(url, email, password, driver_path)
        self.url = url
        self.email = email
        self.password = password
        self.driver_path = driver_path

    def wait(self, seconds):
        return WebDriverWait(self.driver, seconds)

    def close(self):
        self.driver.close()
        self.driver = None

    def quit(self):
        self.driver.quit()
        self.driver = None    

    def login(self):
        
        driver_exe = '.domain\\chromedriver.exe'
        credencials = '.\\config\\credenciales.xlsx'

        print('Entrando en la funcion login...')
        print('----------------------------------------------------------------------')
        
        #Seteo variables
        email = self.email
        url = self.url
        driver_path = self.driver_path
        password  = self.password
        
        options = webdriver.ChromeOptions()
            
        self.driver = webdriver.Chrome(driver_path, options=options)
        self.driver.get(url)
        self.driver.maximize_window()
        self.driver.implicitly_wait(40)

        # Controlar evento alert() Notificacion
        try:
            alert = WebDriverWait(self.driver, 35).until(EC.alert_is_present())
            alert = Alert(self.driver)
            alert.accept()
                    
        except:
            print("No se encontró ninguna alerta.")  

        # Seleccionar codigo de servicio
        intentos = 0
        servicio = True
        while (servicio):
            try:
                print('Try en la funcion codigo de servicio...', intentos)
                print('----------------------------------------------------------------------')
                intentos += 1
                element_codigo= self.driver.find_element(By.ID, 'codigo')
                element_codigo.click()
                element_codigo.send_keys('200-7')
                servicio = False
            except:    
                print('Exception en la funcion codigo de servicio')
                print('----------------------------------------------------------------------')
                servicio = intentos <= 3                

        # Hacer click para ingresar
        intentos = 0
        ingreso = True
        while (ingreso):
            try:
                print('Try en la funcion click para ingresar..', intentos)
                print('----------------------------------------------------------------------')
                intentos += 1
                time.sleep(5) 
                button_element = self.driver.find_element(By.CLASS_NAME, "btn-u")
                if not button_element.is_displayed():
                    action_chains = ActionChains(self.driver)
                    action_chains.move_to_element(button_element).perform()
                button_element.click()
                ingreso = False   
            except:
                print('Exception en la funcion click para ingresar')
                print('----------------------------------------------------------------------')
                ingreso = intentos <= 3      
         
    def scrapping_codiner(self):

        print('Entrando en la funcion Scrapping...')
        print('----------------------------------------------------------------------')
        
        folder_path_config = './config/'
        
        # Especifica la ruta de tu archivo Excel
        excel_file = folder_path_config + "clientes.xlsx"

        # Especifica el nombre de la hoja en la que se encuentran los datos
        hoja_excel = "Hoja1"

        # Carga los datos de Excel en un DataFrame
        df = pd.read_excel(excel_file, sheet_name=hoja_excel)
        print('Dataframe ', df)
        print('----------------------------------------------------------------------')
        
        #Esperamos detectar iframe para poder obtener los datos de la tabla
        wait = WebDriverWait(self.driver, 10)
        iframe_element = wait.until(EC.presence_of_element_located((By.NAME, "window")))
        
        #Hacemos el cambio de iframe    
        self.driver.switch_to.frame(iframe_element)
        
        #Hacemos diccionario para completar con url y fecha
        dicc = {}
        
        #Iteramos para obtener las url de descarga y la fecha
        i=1
        while i < 6:
        
            try:
                time.sleep(5)
                #Buscamos el elemento que contiene la url
                td_element = self.driver.find_element(By.ID, f"td{i}")
                a_element = td_element.find_element(By.TAG_NAME, "a")
                #Guardamos la url y la fecha
                url_a_hacer_clic = a_element.get_attribute("href")
                fecha = a_element.text
                dicc[f"{url_a_hacer_clic}"] = fecha
                time.sleep(5)
            except:
                print('No se han encontrado archivos de descarga')
                print('----------------------------------------------------------------------')
                    
            i += 2

        #Iteramos sobre diccionario para descargar pdf
        for url, date in dicc.items():
            
            #Separamos la fecha para tener mes y año
            mes,año = date.split()
            #Ejecutamos la url en otra ventana
            self.driver.execute_script("window.open(arguments[0]);", url)
        
            # Esperar a que se abra la ventana emergente
            intentos = 0
            reintentar_ventana = True
            while (reintentar_ventana):
                    try:
                        # Obtiene el identificador de la ventana actual
                        current_window = self.driver.current_window_handle
                        print('Ventana principal: ', current_window)
                        print('----------------------------------------------------------------------')
                        print('Try en la funcion manejo de ventanas abiertas...', intentos)
                        print('----------------------------------------------------------------------')
                        intentos += 1
                        window_handles_all = self.driver.window_handles
                        reintentar_ventana = False
                            
                    except:    
                        print('Exception en la funcion manejo de ventanas abiertas')
                        print('----------------------------------------------------------------------')
                        print('----------------------------------------------------------------------')
                        time.sleep(60) #espera para que cargue ventana emergente
                                    
                        reintentar_ventana = intentos <= 3
                                
            # Obtiene los identificadores de las ventanas abiertas
            self.driver.implicitly_wait(35)
            print('Ventanas abiertas: ', window_handles_all, len(window_handles_all))
            print('----------------------------------------------------------------------')            
                                
            # Cambiar al manejo de la ventana emergente
            for window_handle in window_handles_all:
                    if window_handle != current_window:
                        self.driver.switch_to.window(window_handle)
                        print('Ventana emergente: ', window_handle)
                        print('----------------------------------------------------------------------')
                        break

            # Esperar hasta que el elemento esté presente en la página
            self.driver.implicitly_wait(35)
                            
            try:
                # Obtener la URL de la ventana emergente
                ventana_emergente_url = self.driver.current_url
                print("URL de la ventana emergente:", ventana_emergente_url)
                print('----------------------------------------------------------------------')

                # Realizar una solicitud GET para obtener la data binaria del documento
                response = requests.get(ventana_emergente_url, stream=True)

                #Expresion regular para extraer 
                string_split = ventana_emergente_url
                regexp_split = re.split(pattern = r"[>: \: \ \_ \: \< ]", string = string_split)
                print('Resultante expresion regular con split: ', regexp_split[3][:-4])
                print('--------------------------------------------------------------------------')
                bill_number = regexp_split[3][:-4]

                # Cruce datos faltantes para ontener
                for index, row in df.iterrows():
                    
                    df_nro_cliente = df.loc[index, 'nro_cliente']
                    df_sucursal = df.loc[index, 'sucursal']
                    print('Nro. Cliente: ', df_nro_cliente, 'Sucursal: ', df_sucursal)
                    print('--------------------------------------------------------------------------')

                # Obtener el nombre del archivo a partir de los datos del proceso de descarga
                folder_path = './input/'
                file_name = folder_path + str(df_nro_cliente) +"_"+ str(bill_number)+"_"+ str(f'{año}')+".pdf" 

                # # Obtener el nombre del archivo a partir de los datos del proceso de descarga
                # folder_path = './input/'
                # file_name = folder_path + str('Codiner') +"_"+ str(f'{mes}')+"_"+ str(f'{año}')+".pdf" 

                # Guardar la data binaria en un archivo PDF
                with open(file_name, 'wb') as file:
                    response.raw.decode_content = True
                    shutil.copyfileobj(response.raw, file)
                    print("Guardando archivo:", file_name)
                    print('----------------------------------------------------------------------')
                                    
            except:    
                print("No se encontró el elemento con el id especificado...")
                print('----------------------------------------------------------------------')
                                
                count += 1
                print('Conteo de documentos: ', count)
                print('----------------------------------------------------------------------')                

            # Cerrar la ventana emergente
            time.sleep(5)
            self.driver.close()
            time.sleep(15)                            

            # Cambiar de nuevo al manejo de ventana principal
            self.driver.switch_to.window(current_window)
            print('Cual ventana es: ', current_window)
            print('----------------------------------------------------------------------')
               
            #Hacemos una pequeña pausa antes de pasar al siguiente archivo         
            time.sleep(5)
                
            print('pasamos al siguiente archivo')

    def diccionario(self,mes_texto):
        dic = {'Ene':'01','Feb':'01','Mar':'01','Abr':'01','May':'01','Jun':'01','Jul':'01',
               'Ago':'01','Sep':'01','Oct':'01','Nov':'01','Dic':'01'}

        mes_oficial = dic[mes_texto]
        
        return mes_oficial

    def archivos(self):
            
        folder_path = './input/'
        output_path = './output/'
        
        #Revisamos si hay archivos pdf en la carpeta input
        archivos_pdf = glob.glob(os.path.join(folder_path, '*.pdf'))

        #Si no encuentra archivos es porque no se realizo la ejecucion correcta y hay que mandar mail
        if not archivos_pdf:
            print(f'No se encontraron archivos PDF en la carpeta "{folder_path}".')
        else:
            #Si encuentra me entregara todos los documentos con los que trabajaremos
            print(f'Se encontraron los siguientes archivos PDF en la carpeta "{folder_path}":')
            
            for archivo in archivos_pdf:
                
                nombre_oficial = archivo.replace('./input','')
                            
                with fitz.open(archivo) as pdf_documento:
                    texto_completo = ''

                    for pagina_num in range(pdf_documento.page_count):
                        pagina = pdf_documento.load_page(pagina_num)
                        texto_completo += pagina.get_text()
                    
                    lista_limpia_2 = [elemento.strip() for elemento in texto_completo.split('\n')]

                    lista_limpia = [re.sub(r'\s+', ' ', elemento).strip() for elemento in lista_limpia_2]
                        
                    #Posicion 0.0
                    texto_a_verificar = 'R.U.T.'
                    posicion_0_0 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_0_0 = idx
                            tipo_documento_bruto = lista_limpia[posicion_0_0+1]
                            partes = tipo_documento_bruto.split(" ")
                            tipo_documento = partes[0]
                            n_documento_bruto = lista_limpia[posicion_0_0+2]
                            n_documento = n_documento_bruto.replace('Nº ','')
                            break
    
                    #Posicion 0.1
                    elemento_a_buscar = 'Nº CLIENTE'
                    try:
                        n_cliente_bruto = lista_limpia.index(elemento_a_buscar)
                        n_cliente = n_cliente_bruto[posicion_10+1]                  
                    except:
                        n_cliente = ''
                        
                    
                    #Posicion 1
                    texto_a_verificar = 'Fecha de emisión:'
                    posicion_2 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_2 = idx
                            fecha_emision_bruto = lista_limpia[posicion_2]
                            indice_dos_puntos = fecha_emision_bruto.index(":") + 1
                            fecha = fecha_emision_bruto[indice_dos_puntos:].strip()
                            partes = fecha.split()
                            dia = partes[0]
                            mes_match = partes[1]
                            año = partes[2]
                            mes = self.diccionario(mes_match)
                            fecha_emision = dia+'.'+mes+'.'+año
                            break

                    #Posicion 2
                    texto_a_verificar = 'Sr.(a)'
                    posicion_2 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_2 = idx
                            nombre_cliente_bruto = lista_limpia[posicion_2]
                            nombre_cliente = nombre_cliente_bruto.replace('Sr.(a) ','') 
                            break

                    #Posicion 3
                    texto_a_verificar = 'Dirección de envío:'
                    posicion_3 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_3 = idx
                            direccion_bruto = lista_limpia[posicion_3]
                            direccion = direccion_bruto.replace("Dirección de envío: ", "")
                            break
                    
                    #Posicion 4
                    texto_a_verificar = 'Ruta:'
                    posicion_4 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_4 = idx
                            ruta_bruto = lista_limpia[posicion_4]
                            partes = ruta_bruto.split(':')
                            ruta_b = partes[1].split(' | ')
                            ruta= ruta_b[0].strip()
                            break
                    
                    #Posicion 5
                    texto_a_verificar = 'Subestación:'
                    posicion_5 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_5 = idx
                            subestacion_bruto = lista_limpia[posicion_5]
                            subestacion_bruto_ = subestacion_bruto.split(':')
                            subestacion = subestacion_bruto_[1]
                            break

                    #Posicion 6
                    texto_a_verificar = 'Potencia conectada:'
                    posicion_6 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_6 = idx
                            potencia_conectada_b = lista_limpia[posicion_6]
                            potencia_conectada_b_ = potencia_conectada_b.split(': ')
                            potencia_conectada_total = potencia_conectada_b_[1]
                            potencia_conectada = potencia_conectada_total.replace(' kW','')
                            break
                    
                    #Posicion 7
                    texto_a_verificar = 'Fecha término de tarifa:'
                    posicion_7 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_7 = idx
                            fecha_termino_bruto = lista_limpia[posicion_7]
                            fecha_termino_bruto_ = fecha_termino_bruto.split(': ')
                            fecha_termino = fecha_termino_bruto_[1]
                            break
                    
                    #Posicion 8
                    texto_a_verificar = 'Fecha límite para cambio de tarifa:'
                    posicion_8 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_8 = idx
                            fecha_límite_cambio = lista_limpia[posicion_8+1]
                            break
                    
                    #Posicion 9
                    texto_a_verificar = 'Tipo de tarifa contratada'
                    posicion_9 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_9 = idx
                            tipo_tarifa_bruto = lista_limpia[posicion_9]
                            tipo_tarifa_bruto_ = tipo_tarifa_bruto.split(': ')
                            tipo_tarifa = tipo_tarifa_bruto_[1]
                            tipo_tarifa_homologada = tipo_tarifa.replace(' ','')
                            break
                    
                    #Posicion 10
                    texto_a_verificar = 'Período de Lectura:'
                    posicion_10 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            try:
                                posicion_10 = idx
                                lecturas = lista_limpia[posicion_10]
                                print(lecturas)
                                partes = lecturas.split(' - ')
                                fecha_lectura_anterior_b = partes[0]
                                fecha_lectura_anterior = fecha_lectura_anterior_b.replace('Período de Lectura: ','')
                                fecha_lectura_mes_consumo = partes[1]
                                break
                            except:
                                fecha_lectura_anterior = ''
                                fecha_lectura_mes_consumo = ''              
                    
                    #Posicion 11
                    elemento_a_buscar = 'Constante Consumo medidor'
                    try:
                        posicion_11 = lista_limpia.index(elemento_a_buscar)
                        lectura_actual = lista_limpia[posicion_11+5]
                        lectura_anterior = lista_limpia[posicion_11+6]
                        constante = lista_limpia[posicion_11+8]
                        consumo_medidor_kwh_bruto = lista_limpia[posicion_11+9]
                        consumo_medidor_kwh = consumo_medidor_kwh_bruto.replace(' kWh','')
                        consumo_medidor_kvar_bruto = lista_limpia[posicion_11+18]
                        consumo_medidor_kvar = consumo_medidor_kvar_bruto.replace(' KVarh','')
                        
                    except:
                        print('elemento no se encuentra disponible')
                        fecha_emision = '-'
                    
                    #Posicion 12
                    texto_a_verificar = 'Fecha estimada próxima lectura:'
                    posicion_12 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_12 = idx
                            fecha_prox_lectura_br = lista_limpia[posicion_12]
                            fecha_prox_lectura_br_ = fecha_prox_lectura_br.split(': ')
                            fecha_prox_lectura = fecha_prox_lectura_br_[1]
                            break
                    
                    #Posicion 13
                    texto_a_verificar = 'Demanda horas punta:'
                    posicion_13 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_13 = idx
                            demanda_hora_punta_b = lista_limpia[posicion_13]
                            partes = demanda_hora_punta_b.split(":")
                            demanda_hora_punta = partes[1]
                            break
                    
                    #Posicion 14
                    texto_a_verificar = 'Demanda máxima:'
                    posicion_14 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_14 = idx
                            demanda_maxima_b = lista_limpia[posicion_14]
                            partes = demanda_maxima_b.split(":")
                            demanda_maxima = partes[1]
                            break
                    
                    #Posicion 15
                    texto_a_verificar = 'Administracion del servicio (Cargo fijo mensual)'
                    posicion_15 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_15 = idx
                            try:
                                administracion_servicio_b = lista_limpia[posicion_15+2]
                                administracion_servicio = administracion_servicio_b.replace('.','')
                                break
                            except:
                                print('elemento no se encuentra disponible')
                                administracion_servicio = ''
                                
                    #Posicion 16
                    texto_a_verificar = 'Transporte de Electricidad'
                    posicion_16 = None
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_16 = idx
                            try:
                                cantidad_transporte_electricidad_b = lista_limpia[posicion_16+2]
                                cantidad_transporte_electricidad = cantidad_transporte_electricidad_b.replace('.','')
                                break
                            except:
                                print('elemento no se encuentra disponible')
                                cantidad_transporte_electricidad = ''
                                

                    #Posicion 17
                    texto_a_verificar = 'Electricidad Consumida'
                    posicion_17 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_17 = idx
                            try:
                                electricidad_consumida_b = lista_limpia[posicion_17+2]
                                electricidad_consumida = electricidad_consumida_b.replace('.','')
                                break
                            except:
                                print('elemento no se encuentra disponible')
                                electricidad_consumida = ''
                    
                    #Posicion 18
                    texto_a_verificar = 'Cargo por demanda maxima de potencia suministrada'
                    posicion_18 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_18 = idx
                            cantidad_carga_dem_suminis_b = lista_limpia[posicion_18]
                            try:
                                cantidad_carga_dem_suminis_b_ = re.search(r'\d+,\d+', cantidad_carga_dem_suminis_b)
                                cantidad_carga_dem_suminis = cantidad_carga_dem_suminis_b_.group()
                                valor_carga_dem_suminis_b = lista_limpia[posicion_18+2]
                                valor_carga_dem_suminis = valor_carga_dem_suminis_b.replace('.','')
                                break
                            except:
                                print('elemento no se encuentra disponible')
                                valor_carga_dem_suminis = ''
                                
    
                    #Posicion 19
                    texto_a_verificar = 'Cargo por demanda maxima de potencia en horas de punta'
                    posicion_19 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_19 = idx
                            cantidad_carga_dem_suminis_punta_b = lista_limpia[posicion_19]
                            try:
                                cantidad_carga_dem_suminis_punta_b_ = re.search(r'\d+,\d+', cantidad_carga_dem_suminis_punta_b)
                                cantidad_carga_dem_suminis_punta = cantidad_carga_dem_suminis_punta_b_.group()
                                carga_dem_suminis_punta_b = lista_limpia[posicion_19+2]
                                valor_carga_dem_suminis_punta = carga_dem_suminis_punta_b.replace('.','')
                                break
                            except:
                                print('elemento no se encuentra disponible')
                                valor_carga_dem_suminis_punta = ''

                    #Posicion 20
                    texto_a_verificar = 'Pago de la cuenta fuera de plazo'
                    posicion_20 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_20 = idx
                            try:
                                fuera_plazo_b = lista_limpia[posicion_20+2]
                                fuera_plazo = fuera_plazo_b.replace('.','')
                                break
                            except:
                                print('elemento no se encuentra disponible')
                                fuera_plazo = ''
                     
                    # #Posicion 21
                    # texto_a_verificar = 'Interes por mora'
                    # posicion_21 = None 
                    
                    # for idx, elemento in enumerate(lista_limpia):
                    #     if texto_a_verificar in elemento:
                    #         posicion_21 = idx
                    #         try:
                    #             interes_mora_b = lista_limpia[posicion_21+2]
                    #             interes_mora = interes_mora_b.replace('.','')
                    #             break
                    #         except:
                    #             print('elemento no se encuentra disponible')
                    #             interes_mora = ''

                    #Posicion 23
                    texto_a_verificar = 'Cargo Fondo Estabilizacion Ley 21472 (Exento)'
                    posicion_23 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_23 = idx
                            try:
                                fondo_estabilizacion_b = lista_limpia[posicion_23+2]
                                fondo_estabilizacion = fondo_estabilizacion_b.replace('.','')
                                break
                            except:
                                print('elemento no se encuentra disponible')
                                fondo_estabilizacion = ''
                            
                    #Posicion 24
                    texto_a_verificar = 'Diferencia Ajuste Sencillo'
                    posicion_24 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_24 = idx
                            try:
                                suma_sencillos = lista_limpia[posicion_24+2]
                                break
                            except:
                                print('elemento no se encuentra disponible')
                                suma_sencillos = ''
                    
                    #Posicion 25
                    try:
                        fecha_vencimiento_b = lista_limpia[0]
                        partes_2 = fecha_vencimiento_b.split()
                        dia_ven = partes_2[0]
                        mes_ven = partes_2[1]

                        mes_ven = self.diccionario(partes_2[1])
                        año_ven = partes_2[2]
                        fecha_vencimiento = dia_ven+'-'+mes_ven+'-'+año_ven
                    except:
                        print('elemento no se encuentra disponible')
                        fecha_vencimiento = ''

                    #Posicion 26   
                    #NETO
                    elemento_a_buscar = 'Neto'
                    try:
                        posicion_26 = lista_limpia.index(elemento_a_buscar)
                        neto_bruto = lista_limpia[posicion_26+2]
                        neto = neto_bruto.replace('.','')    
                    except:
                        print('elemento no se encuentra disponible')
                        neto = ''
                    
                    #Posicion 27
                    elemento_a_buscar = 'Saldo Anterior'
                    try:
                        posicion_27 = lista_limpia.index(elemento_a_buscar)
                        saldo_anterior_bruto = lista_limpia[posicion_27+2]
                        saldo_anterior = saldo_anterior_bruto.replace('.','')    
                    except:
                        print('elemento no se encuentra disponible')
                        saldo_anterior = ''
                    
                    #Posicion 28
                    elemento_a_buscar = 'Total a pagar'
                    try:
                        posicion_28 = lista_limpia.index(elemento_a_buscar)
                        total_pagar_bruto = lista_limpia[posicion_28+1]
                        total_pagar = total_pagar_bruto.replace('.','').replace('$ ','')    
                    except:
                        print('elemento no se encuentra disponible')
                        total_pagar = ''

                    #Posicion 29 (arreglo #1)
                    texto_a_verificar = 'Transporte de Electricidad'
                    posicion_29 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_29 = idx
                            try:
                                valor_transporte_elec_b = lista_limpia[posicion_29+2]
                                valor_transporte_elec = valor_transporte_elec_b.replace('.','')
                                break
                            except:
                                print('elemento no se encuentra disponible')
                                valor_transporte_elec = ''
                    
                    #Posicion 30 (arreglo #2)
                    elemento_a_buscar = 'Pago de la cuenta fuera de plazo'
                    try:
                        posicion_30 = lista_limpia.index(elemento_a_buscar)
                        total_pagar_fuera_plazo_b = lista_limpia[posicion_30+2]
                        total_pagar_fuera_plazo = total_pagar_fuera_plazo_b.replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        total_pagar_fuera_plazo = ''
                    
                    #Posicion 31 (arreglo #3)
                    elemento_a_buscar = 'Interes por mora'
                    try:
                        posicion_31 = lista_limpia.index(elemento_a_buscar)
                        interes_mora_b = lista_limpia[posicion_31+2]
                        interes_mora = interes_mora_b.replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        interes_mora = ''
                        
                    #Posicion 32 (arreglo #4)
                    texto_a_verificar = 'Cargo Fondo Estabilizacion'
                    posicion_32 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_32 = idx
                            try:
                                valor_fondo_estabilizacion_b= lista_limpia[posicion_32+2]
                                valor_fondo_estabilizacion = valor_fondo_estabilizacion_b.replace('.','')
                                break
                            except:
                                print('elemento no se encuentra disponible')
                                valor_fondo_estabilizacion = ''

                #Cargamos libro excel donde volcaremos los datos
                libro = load_workbook(output_path+'/'+'Formato Planilla.xlsx')
                hoja_electricidad = libro['Electricidad']
                    
                ultima_fila = hoja_electricidad.max_row
                
                #Los datos mas importantes
                hoja_electricidad.cell(row=ultima_fila+1,column=16).value = tipo_documento
                hoja_electricidad.cell(row=ultima_fila+1,column=17).value = int(n_documento)
                
                #Primera tabla traspasada a excel
                hoja_electricidad.cell(row=ultima_fila+1,column=8).value = n_cliente
                hoja_electricidad.cell(row=ultima_fila+1,column=23).value = fecha_emision
                hoja_electricidad.cell(row=ultima_fila+1,column=15).value = nombre_cliente
                hoja_electricidad.cell(row=ultima_fila+1,column=21).value = direccion
                hoja_electricidad.cell(row=ultima_fila+1,column=18).value = ruta
                
                hoja_electricidad.cell(row=ultima_fila+1,column=20).value =subestacion
                hoja_electricidad.cell(row=ultima_fila+1,column=28).value = potencia_conectada
                hoja_electricidad.cell(row=ultima_fila+1,column=25).value = fecha_termino
                hoja_electricidad.cell(row=ultima_fila+1,column=26).value = fecha_límite_cambio
                hoja_electricidad.cell(row=ultima_fila+1,column=10).value = tipo_tarifa
                hoja_electricidad.cell(row=ultima_fila+1,column=11).value = tipo_tarifa_homologada
                
                hoja_electricidad.cell(row=ultima_fila+1,column=14).value = fecha_lectura_anterior
                hoja_electricidad.cell(row=ultima_fila+1,column=13).value = fecha_lectura_mes_consumo
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=30).value = int(lectura_actual)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=30).value = lectura_actual
                
                try:    
                    hoja_electricidad.cell(row=ultima_fila+1,column=31).value = int(lectura_anterior)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=31).value = lectura_anterior
                    
                hoja_electricidad.cell(row=ultima_fila+1,column=27).value = fecha_prox_lectura
                hoja_electricidad.cell(row=ultima_fila+1,column=32).value = int(constante)
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=33).value = int(consumo_medidor_kwh)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=33).value = consumo_medidor_kwh
                    
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=49).value = int(consumo_medidor_kvar)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=49).value = consumo_medidor_kvar
                hoja_electricidad.cell(row=ultima_fila+1,column=45).value = demanda_hora_punta

                hoja_electricidad.cell(row=ultima_fila+1,column=43).value = demanda_maxima
                
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=56).value = int(administracion_servicio)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=56).value = administracion_servicio
                    
                try:    
                    hoja_electricidad.cell(row=ultima_fila+1,column=57).value = int(cantidad_transporte_electricidad)
                except:    
                    hoja_electricidad.cell(row=ultima_fila+1,column=57).value = cantidad_transporte_electricidad
                    
                try:    
                    hoja_electricidad.cell(row=ultima_fila+1,column=59).value = int(electricidad_consumida)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=59).value = electricidad_consumida
                    
                try:    
                    hoja_electricidad.cell(row=ultima_fila+1,column=44).value = int(cantidad_carga_dem_suminis)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=44).value = cantidad_carga_dem_suminis
                
                try:    
                    hoja_electricidad.cell(row=ultima_fila+1,column=63).value = int(valor_carga_dem_suminis)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=63).value = valor_carga_dem_suminis

                hoja_electricidad.cell(row=ultima_fila+1,column=46).value = cantidad_carga_dem_suminis_punta
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=64).value = int(valor_carga_dem_suminis_punta)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=64).value = valor_carga_dem_suminis_punta
                
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=74).value = int(suma_sencillos)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=74).value = suma_sencillos
                    
                hoja_electricidad.cell(row=ultima_fila+1,column=24).value = fecha_vencimiento
                hoja_electricidad.cell(row=ultima_fila+1,column=76).value = int(neto)
                hoja_electricidad.cell(row=ultima_fila+1,column=79).value = int(saldo_anterior)
                hoja_electricidad.cell(row=ultima_fila+1,column=81).value = int(total_pagar)

                #Arreglo 1 
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=110).value = int(valor_transporte_elec)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=110).value = valor_transporte_elec
                
                #Arreglo 2
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=92).value = int(total_pagar_fuera_plazo)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=92).value = total_pagar_fuera_plazo

                #Arreglo 3
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=93).value = int(interes_mora)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=93).value = interes_mora               

                #Arreglo 4
                try:
                    hoja_electricidad.cell(row=ultima_fila+1,column=108).value = int(valor_fondo_estabilizacion)
                except:
                    hoja_electricidad.cell(row=ultima_fila+1,column=108).value = valor_fondo_estabilizacion

                libro.save(output_path+'/'+'Formato Planilla.xlsx')
                    
                # #Copiamos el archivo a la carpeta outpu con el nombre que corresponde
                # shutil.copy(archivo, output_path+nombre_oficial)
                # print('-----')

        # #Obtenemos los archivos de la carpeta input
        # archivos_en_carpeta = os.listdir(folder_path)

        # # Iterar sobre los archivos y eliminarlos
        # for archivo in archivos_en_carpeta:
        #     ruta_archivo = os.path.join(folder_path, archivo)
        #     if os.path.isfile(ruta_archivo):
        #         os.remove(ruta_archivo)